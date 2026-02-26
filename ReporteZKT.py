def generar_reporte(archivo_excel, ruta_salida):

    import pandas as pd
    import os
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Leer archivo
    df = pd.read_excel(archivo_excel, header=2)

    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True)
    df["Hora"] = pd.to_datetime(df["Hora"], format="mixed").dt.time

    df["FechaHora"] = pd.to_datetime(
        df["Fecha"].astype(str) + " " + df["Hora"].astype(str)
    )

    def obtener_horario(fecha):
        dia = fecha.weekday()
        if dia <= 3:
            return "08:30", "18:00"
        elif dia == 4:
            return "08:00", "17:00"
        elif dia == 5:
            return "09:00", "13:00"
        else:
            return None, None

    reporte = []

    for (id_user, fecha), grupo in df.groupby(
        ["Employee ID", df["Fecha"].dt.date]
    ):

        nombre = grupo["Nombres"].iloc[0]

        entrada = grupo["FechaHora"].min()
        salida = grupo["FechaHora"].max()

        hora_ent, hora_sal = obtener_horario(pd.to_datetime(fecha))

        if hora_ent is None:
            continue

        hora_ent = pd.to_datetime(f"{fecha} {hora_ent}")
        hora_sal = pd.to_datetime(f"{fecha} {hora_sal}")

        estado = "OK"

        if entrada > hora_ent + pd.Timedelta(minutes=10):
            estado = "Llegada tarde"
        elif salida < hora_sal - pd.Timedelta(minutes=120):
            estado = "Salida temprano"
        elif salida > hora_sal + pd.Timedelta(minutes=30):
            estado = "Salió tarde"

        if len(grupo) == 1:
            estado = "Marcación incompleta"

        reporte.append([
            id_user,
            nombre,
            fecha,
            entrada.time(),
            salida.time(),
            estado
        ])

    df_reporte = pd.DataFrame(
        reporte,
        columns=["ID", "Nombre", "Fecha", "Entrada", "Salida", "Estado"]
    )

    # Guardar en Descargas
    archivo_salida = os.path.join(
        ruta_salida,
        "Informe_Talento_Humano.xlsx"
    )

    df_reporte.to_excel(archivo_salida, index=False)

    wb = load_workbook(archivo_salida)
    ws = wb.active

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    for fila in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        celda = fila[0]
        if celda.value == "OK":
            celda.fill = verde
        elif celda.value == "Llegada tarde":
            celda.fill = rojo
        elif celda.value == "Salida temprano":
            celda.fill = amarillo
        elif celda.value == "Salió tarde":
            celda.fill = azul
        elif celda.value == "Marcación incompleta":
            celda.fill = rojo

    wb.save(archivo_salida)

    os.startfile(archivo_salida)