import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─── Configuración de página ─────────────────────────────────────────────────
st.set_page_config(
    page_title="Control de Asistencia",
    page_icon="🕐",
    layout="wide"
)

# ─── Estilos ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600&family=DM+Mono&display=swap');

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

    .stApp { background-color: #F7F8FA; }

    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
        color: white;
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
    }
    .main-header h1 { margin: 0; font-size: 1.8rem; font-weight: 600; letter-spacing: -0.5px; }
    .main-header p { margin: 0.4rem 0 0; opacity: 0.65; font-size: 0.9rem; }

    .metric-card {
        background: white;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        border: 1px solid #E8EAF0;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    }
    .metric-label { font-size: 0.78rem; color: #888; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }
    .metric-value { font-size: 2rem; font-weight: 600; color: #1a1a2e; font-family: 'DM Mono', monospace; }

    .status-ok       { background: #C6EFCE; color: #276221; padding: 3px 10px; border-radius: 20px; font-size: 0.82rem; font-weight: 500; }
    .status-tarde    { background: #FFC7CE; color: #9C0006; padding: 3px 10px; border-radius: 20px; font-size: 0.82rem; font-weight: 500; }
    .status-temprano { background: #FFEB9C; color: #7D6608; padding: 3px 10px; border-radius: 20px; font-size: 0.82rem; font-weight: 500; }
    .status-salida   { background: #BDD7EE; color: #1F497D; padding: 3px 10px; border-radius: 20px; font-size: 0.82rem; font-weight: 500; }
    .status-inc      { background: #FFC7CE; color: #9C0006; padding: 3px 10px; border-radius: 20px; font-size: 0.82rem; font-weight: 500; }

    div[data-testid="stFileUploader"] {
        border: 2px dashed #CBD0E0;
        border-radius: 12px;
        background: white;
        padding: 1rem;
    }
    .stButton > button {
        background: #0f3460;
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.6rem 2rem;
        font-family: 'DM Sans', sans-serif;
        font-weight: 500;
        width: 100%;
        transition: background 0.2s;
    }
    .stButton > button:hover { background: #16213e; }

    .download-btn > button {
        background: #276221 !important;
    }
    .download-btn > button:hover { background: #1a4016 !important; }
</style>
""", unsafe_allow_html=True)


# ─── Lógica principal ─────────────────────────────────────────────────────────

def obtener_horario(fecha):
    dia = fecha.weekday()
    if dia <= 3:
        return "08:30", "18:00"
    elif dia == 4:
        return "08:00", "17:00"
    elif dia == 5:
        return "09:00", "13:00"
    else:
        return None, None


def generar_reporte(archivo_excel):
    df = pd.read_excel(archivo_excel, header=2)
    df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True)
    df["Hora"] = pd.to_datetime(df["Hora"], format="mixed").dt.time
    df["FechaHora"] = pd.to_datetime(df["Fecha"].astype(str) + " " + df["Hora"].astype(str))

    reporte = []
    for (id_user, fecha), grupo in df.groupby(["Employee ID", df["Fecha"].dt.date]):
        nombre = grupo["Nombres"].iloc[0]
        entrada = grupo["FechaHora"].min()
        salida = grupo["FechaHora"].max()
        hora_ent, hora_sal = obtener_horario(pd.to_datetime(fecha))

        if hora_ent is None:
            continue

        hora_ent = pd.to_datetime(f"{fecha} {hora_ent}")
        hora_sal = pd.to_datetime(f"{fecha} {hora_sal}")

        estado = "OK"
        if len(grupo) == 1:
            estado = "Marcación incompleta"
        elif entrada > hora_ent + pd.Timedelta(minutes=10):
            estado = "Llegada tarde"
        elif salida < hora_sal - pd.Timedelta(minutes=120):
            estado = "Salida temprano"
        elif salida > hora_sal + pd.Timedelta(minutes=30):
            estado = "Salió tarde"

        reporte.append([id_user, nombre, fecha, entrada.time(), salida.time(), estado])

    return pd.DataFrame(reporte, columns=["ID", "Nombre", "Fecha", "Entrada", "Salida", "Estado"])


def exportar_excel(df_reporte):
    buffer = io.BytesIO()
    df_reporte.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    colores = {
        "OK":                   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Llegada tarde":        PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Salida temprano":      PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Salió tarde":          PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "Marcación incompleta": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for fila in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        celda = fila[0]
        if celda.value in colores:
            celda.fill = colores[celda.value]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def badge_estado(estado):
    clases = {
        "OK": "status-ok",
        "Llegada tarde": "status-tarde",
        "Salida temprano": "status-temprano",
        "Salió tarde": "status-salida",
        "Marcación incompleta": "status-inc",
    }
    cls = clases.get(estado, "status-ok")
    return f'<span class="{cls}">{estado}</span>'


# ─── UI ───────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="main-header">
    <h1>🕐 Control de Asistencia</h1>
    <p>Talento Humano · Generación de informes automáticos</p>
</div>
""", unsafe_allow_html=True)

col_upload, col_info = st.columns([2, 1])

with col_upload:
    archivo = st.file_uploader("Subir archivo Excel de marcaciones", type=["xlsx", "xls"])

with col_info:
    st.markdown("""
    **Horarios configurados**
    - Lunes – Jueves: 08:30 → 18:00
    - Viernes: 08:00 → 17:00
    - Sábado: 09:00 → 13:00
    - Domingo: No laboral
    """)

if archivo:
    if st.button("⚙️ Generar Reporte"):
        with st.spinner("Procesando marcaciones..."):
            try:
                df_reporte = generar_reporte(archivo)
                st.session_state["reporte"] = df_reporte
                st.success(f"✅ Reporte generado — {len(df_reporte)} registros procesados")
            except Exception as e:
                st.error(f"Error al procesar el archivo: {e}")

if "reporte" in st.session_state:
    df = st.session_state["reporte"]

    # ── Métricas ──────────────────────────────────────────────────────────────
    st.markdown("### Resumen")
    conteos = df["Estado"].value_counts()

    m1, m2, m3, m4, m5 = st.columns(5)
    metricas = [
        (m1, "✅ OK",                   conteos.get("OK", 0)),
        (m2, "🔴 Llegada tarde",        conteos.get("Llegada tarde", 0)),
        (m3, "🟡 Salida temprano",      conteos.get("Salida temprano", 0)),
        (m4, "🔵 Salió tarde",          conteos.get("Salió tarde", 0)),
        (m5, "⚠️ Marc. incompleta",     conteos.get("Marcación incompleta", 0)),
    ]
    for col, label, val in metricas:
        with col:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">{label}</div>
                <div class="metric-value">{val}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Filtros ───────────────────────────────────────────────────────────────
    f1, f2 = st.columns(2)
    with f1:
        estados_filtro = st.multiselect(
            "Filtrar por estado",
            options=df["Estado"].unique().tolist(),
            default=df["Estado"].unique().tolist()
        )
    with f2:
        buscar = st.text_input("Buscar por nombre o ID", placeholder="Ej: Juan García...")

    df_filtrado = df[df["Estado"].isin(estados_filtro)]
    if buscar:
        mask = (
            df_filtrado["Nombre"].str.contains(buscar, case=False, na=False) |
            df_filtrado["ID"].astype(str).str.contains(buscar, case=False, na=False)
        )
        df_filtrado = df_filtrado[mask]

    # ── Tabla con badges ──────────────────────────────────────────────────────
    st.markdown("### Detalle de registros")

    df_html = df_filtrado.copy()
    df_html["Estado"] = df_html["Estado"].apply(badge_estado)
    df_html["Fecha"] = df_html["Fecha"].astype(str)
    df_html["Entrada"] = df_html["Entrada"].astype(str).str[:5]
    df_html["Salida"] = df_html["Salida"].astype(str).str[:5]

    tabla_html = df_html.to_html(escape=False, index=False, classes="dataframe")
    st.markdown(f"""
    <style>
        table.dataframe {{
            width: 100%; border-collapse: collapse;
            background: white; border-radius: 10px; overflow: hidden;
            box-shadow: 0 1px 4px rgba(0,0,0,0.06);
            font-size: 0.88rem;
        }}
        table.dataframe th {{
            background: #1a1a2e; color: white;
            padding: 10px 14px; text-align: left; font-weight: 500;
            font-size: 0.8rem; letter-spacing: 0.3px;
        }}
        table.dataframe td {{
            padding: 9px 14px; border-bottom: 1px solid #F0F2F7;
            color: #333;
        }}
        table.dataframe tr:last-child td {{ border-bottom: none; }}
        table.dataframe tr:hover td {{ background: #F7F8FA; }}
    </style>
    {tabla_html}
    """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Descarga ──────────────────────────────────────────────────────────────
    excel_bytes = exportar_excel(df)
    st.markdown('<div class="download-btn">', unsafe_allow_html=True)
    st.download_button(
        label="📥 Descargar Informe Excel",
        data=excel_bytes,
        file_name="Informe_Talento_Humano.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.markdown('</div>', unsafe_allow_html=True)
